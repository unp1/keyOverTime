#!/usr/bin/env python3
"""Write the comparison as a workbook: one tidy sheet plus one pivot sheet per metric.

  gen_xlsx.py out/tidy.csv -o out/keyOverTime.xlsx [--corpus corpus.csv]

No charts. A native spreadsheet chart is re-rendered by whatever opens the file, and Excel, Numbers
and LibreOffice disagree about log axes, per-series end labels and exact series colours -- the three
things these figures depend on. The charts live in the PDF, which keeps them as drawn; this file is
for reading, sorting and deriving from the numbers.

Each metric sheet is versions across, proofs down, with a ratio column of the newest version against
the oldest and a colour scale over the per-cell ratio to the baseline. Cells with no measurement are
not left blank: a proof deliberately excluded from a version says so, so an exclusion never reads as
missing data.
"""
import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from csvio import read_rows  # noqa: E402

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from gen_report import COLUMNS, METRICS, PROOFS, build

HEAD_FILL = PatternFill("solid", fgColor="F4F3F0")
HEAD_FONT = Font(bold=True, color="52514E", size=10)
TITLE_FONT = Font(bold=True, size=13, color="0B0B0B")
NOTE_FONT = Font(size=9, color="7C7A74")
EXCL_FONT = Font(size=9, color="7C7A74", italic=True)
MT_FILL = PatternFill("solid", fgColor="FAF9F6")   # the goal-parallel columns
FLAG_FONT = Font(size=10, color="D03B3B", bold=True)
THIN = Side(style="thin", color="DBDAD4")
BORDER = Border(bottom=THIN)


def read_corpus(path):
    """{label: set(configs) or None} -- None meaning the proof runs on every version."""
    only = {}
    try:
        for f in read_rows(path, min_fields=2, pad=6):
            only[f[1]] = set(f[5].replace(",", " ").split()) if f[5].strip() else None
    except OSError as e:
        print(f"warning: {e}", file=sys.stderr)
    return only


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tidy")
    ap.add_argument("-o", "--out", required=True)
    ap.add_argument("--corpus", default=None)
    args = ap.parse_args()

    with open(args.tidy) as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        print("no rows", file=sys.stderr)
        return 1
    only = read_corpus(args.corpus) if args.corpus else {}

    idx = {(r["mode"], r["label"], r["config"]): r for r in rows}
    _values, status = build(rows)

    wb = Workbook()

    # ---------------------------------------------------------------- tidy sheet
    ws = wb.active
    ws.title = "data"
    fields = list(rows[0].keys())
    ws.append(fields)
    for c in range(1, len(fields) + 1):
        ws.cell(1, c).fill = HEAD_FILL
        ws.cell(1, c).font = HEAD_FONT
    for r in rows:
        ws.append([_num(r[f]) for f in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
    for c, f in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(18, len(f) + 4))

    # ---------------------------------------------------------------- one sheet per metric
    for mid, title, unit, blurb in METRICS:
        if not any(r.get(mid) not in ("", None) for r in rows):
            continue
        ws = wb.create_sheet(_sheet_name(title))
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"{blurb}  Unit: {unit}."
        ws["A2"].font = NOTE_FONT
        _block(ws, 4, mid, idx, only)
        ws.column_dimensions["A"].width = 24
        for c in range(2, len(COLUMNS) + 4):
            ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(args.out)
    print(f"wrote {args.out}", file=sys.stderr)
    return 0


def _block(ws, row, metric, idx, only):
    """One columns-across, proofs-down table; returns the next free row.

    A column is a (version, prover mode) pair, matching the charts: the goal-parallel runs sit
    beside the single-threaded run of the same build rather than in a separate table, so a reader
    comparing them does not have to scroll between two blocks.
    """
    header = ["proof"] + [n + (" " + q if q else "") for _c, _m, n, q, _d in COLUMNS] \
        + ["last / first"]
    for c, h in enumerate(header, 1):
        cell = ws.cell(row, c, h)
        cell.fill, cell.font, cell.border = HEAD_FILL, HEAD_FONT, BORDER
        cell.alignment = Alignment(horizontal="right" if c > 1 else "left", wrap_text=True)
    head_row = row
    row += 1

    first_data = row
    for pid, name, _light, _dark in PROOFS:
        vals = []
        ws.cell(row, 1, name).border = BORDER
        for c, (cfg, mode, cname, qual, _note) in enumerate(COLUMNS, 2):
            rec = idx.get((mode, pid, cfg))
            cell = ws.cell(row, c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            if mode == "mt":
                cell.fill = MT_FILL
            allowed = only.get(pid)
            if rec is None:
                if allowed is not None and cfg not in allowed:
                    cell.value = "excluded"
                    cell.font = EXCL_FONT
                    cell.comment = Comment(
                        f"{name} is measured on {', '.join(sorted(allowed))} only; on this version "
                        f"it does not close, so its numbers would not be comparable.", "key-bench")
                else:
                    cell.value = "\u2014"
                    cell.font = EXCL_FONT
                vals.append(None)
                continue
            v = _num(rec.get(metric, ""))
            cell.value = v
            cell.number_format = "#,##0" if isinstance(v, (int, float)) and v >= 100 else "0.000"
            if rec.get("status") not in ("ok", "", None):
                cell.font = FLAG_FONT
                cell.comment = Comment(
                    f"status {rec['status']}: {rec.get('openGoals', '?')} open goal(s). "
                    "Did not close, so this time is not comparable with a closed proof.",
                    "key-bench")
            vals.append(v if isinstance(v, (int, float)) else None)

        c = len(COLUMNS) + 2
        base = vals[0]
        last = next((v for v in reversed(vals) if v is not None), None)
        cell = ws.cell(row, c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right")
        if base and last is not None:
            cell.value = last / base
            cell.number_format = "0.00\u00d7"
        else:
            cell.value = "\u2014"
            cell.font = EXCL_FONT
        row += 1

    # A colour scale across the value columns, so a row reads as a trajectory at a glance.
    rng = (f"{get_column_letter(2)}{first_data}:"
           f"{get_column_letter(len(COLUMNS) + 1)}{row - 1}")
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="CDE2FB",
        mid_type="percentile", mid_value=50, mid_color="F0EFEC",
        end_type="max", end_color="EB6834"))
    ws.freeze_panes = ws.cell(head_row + 1, 2)
    return row + 2


def _sheet_name(title):
    bad = set(r"[]:*?/\\")
    name = "".join(ch for ch in title if ch not in bad)
    return name[:31]


def _num(v):
    if v in ("", None):
        return ""
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return v


if __name__ == "__main__":
    sys.exit(main())
